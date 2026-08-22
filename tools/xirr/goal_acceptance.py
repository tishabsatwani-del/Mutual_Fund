"""Acceptance tests for the Plan my goal tab.

The sheet sums a year of instalments in closed form, because Excel cannot loop.
This file checks it the other way -- month by month, in Python -- so agreement
means two different methods reached the same number, not that one method agrees
with itself.
"""
import csv
import glob
import os
import shutil
import subprocess
import sys

from openpyxl import load_workbook

SRC = sys.argv[1]
WORK = "goal_run"
PROFILE = "file:///tmp/lo-profile-xirr"
GOAL_SHEET = 3          # Start here, My investments, Plan my goal, ...

failures = []


def report(name, ok, detail=""):
    print(f"{'PASS' if ok else 'FAIL'}  {name}{('   -- ' + detail) if detail else ''}")
    if not ok:
        failures.append(name)


def close(name, actual, expected, tol=0.75):
    """Money, so agreement to within a rupee is agreement."""
    ok = actual is not None and abs(actual - expected) <= tol
    report(name, ok, f"sheet {actual}, calculated {expected:.2f}" if actual is not None else "no value")


# ---------------------------------------------------------------- month by month
def fv_sip(monthly, annual_rate, years, step_up=0.0):
    i = (1 + annual_rate) ** (1 / 12) - 1
    amount, total = monthly, 0.0
    for m in range(int(round(years * 12))):
        if m > 0 and m % 12 == 0:
            amount *= (1 + step_up)
        total = (total + amount) * (1 + i)      # paid at the start of the month
    return total


def fv_lump(present, annual_rate, years):
    return present * (1 + annual_rate) ** years


def contributions(monthly, years, step_up=0.0):
    amount, total = monthly, 0.0
    for m in range(int(round(years * 12))):
        if m > 0 and m % 12 == 0:
            amount *= (1 + step_up)
        total += amount
    return total


# --------------------------------------------------------------------- the sheet
def run(case, values):
    """Type the inputs in, recalculate outside Excel, read the answers back."""
    wb = load_workbook(SRC)
    ws = wb["Plan my goal"]
    ws.protection.sheet = False
    for cell, value in values.items():
        ws[cell] = value
    # widen the display so the checks read the number, not a rounded rendering
    for cell in ["B11", "B15", "B18", "B19", "B20", "B24", "B25", "B26", "B27"]:
        ws[cell].number_format = "0.00"
    path = os.path.join(WORK, f"{case}.xlsx")
    wb.save(path)

    outdir = os.path.join(WORK, f"out{case}")
    os.makedirs(outdir, exist_ok=True)
    flt = ("csv:Text - txt - csv (StarCalc):44,34,76,1,,0,false,true,true,"
           f"false,false,{GOAL_SHEET}")
    subprocess.run(
        ["soffice", f"-env:UserInstallation={PROFILE}", "--headless",
         "--convert-to", flt, "--outdir", outdir, path],
        check=True, capture_output=True, timeout=300,
    )
    with open(glob.glob(os.path.join(outdir, "*.csv"))[0], newline="", encoding="utf-8") as fh:
        grid = list(csv.reader(fh))

    def cell(ref):
        col = ord(ref[0]) - 65
        row = int(ref[1:]) - 1
        try:
            return grid[row][col]
        except IndexError:
            return ""

    def number(ref):
        raw = cell(ref).replace(",", "").replace("₹", "").strip()
        try:
            return float(raw)
        except ValueError:
            return None

    return cell, number


shutil.rmtree(WORK, ignore_errors=True)
os.makedirs(WORK, exist_ok=True)

# ------------------------------------------------------------------------ case 1
print("\nCase 1 — a lump sum, no monthly investing")
inputs = {"B4": 1000000, "B5": 400000, "B6": 5, "B7": 0, "B8": 8, "B9": 0}
cell, number = run("one", inputs)
expected = fv_lump(400000, 0.08, 5)
close("projected value matches compounding done by hand", number("B11"), expected)
close("the whole projection comes from the existing corpus", number("B18"), expected)
close("nothing comes from monthly investing", number("B19"), 0)
report("the shortfall is stated in words", "Short by" in cell("B14"), cell("B14"))
extra = number("B15")
close("the top-up matches the gap divided by what a rupee a month grows to",
      extra, (1000000 - expected) / fv_sip(1, 0.08, 5, 0))

print("\nCase 1b — feeding that top-up back in must land on the goal")
cell, number = run("oneb", {**inputs, "B7": extra})
close("adding the required top-up reaches the goal", number("B11"), 1000000, 2.0)
report("landing on the goal is reported as covered", "Covered" in cell("B14"), cell("B14"))

# ------------------------------------------------------------------------ case 2
print("\nCase 2 — corpus plus a monthly instalment")
cell, number = run("two", {"B4": 5000000, "B5": 400000, "B6": 15, "B7": 10000, "B8": 10, "B9": 0})
corpus, sip = fv_lump(400000, 0.10, 15), fv_sip(10000, 0.10, 15, 0)
close("the two parts are reported separately", number("B18"), corpus)
close("monthly investing is summed month by month", number("B19"), sip)
close("the projection is the two added", number("B11"), corpus + sip)
close("own money paid in is instalments times months", number("B20"), contributions(10000, 15, 0))

# ------------------------------------------------------------------------ case 3
print("\nCase 3 — with a yearly step-up")
cell, number = run("three", {"B4": 5000000, "B5": 0, "B6": 20, "B7": 5000, "B8": 12, "B9": 10})
close("a stepped-up instalment compounds correctly", number("B11"), fv_sip(5000, 0.12, 20, 0.10))
close("own money paid in accounts for the step-up", number("B20"), contributions(5000, 20, 0.10))

# --------------------------------------------------------------------- scenarios
print("\nScenario rows")
cell, number = run("four", {"B4": 5000000, "B5": 400000, "B6": 15, "B7": 10000, "B8": 10, "B9": 0})
base = fv_lump(400000, 0.10, 15)
close("carry on as you are", number("B24"), base + fv_sip(10000, 0.10, 15, 0))
close("add 2,000 a month", number("B25"), base + fv_sip(12000, 0.10, 15, 0))
close("add 5,000 a month", number("B26"), base + fv_sip(15000, 0.10, 15, 0))
close("same amount, raised 10% a year", number("B27"), base + fv_sip(10000, 0.10, 15, 0.10))
report("a scenario says where it lands against the goal",
       "short by" in cell("C24") or "covered" in cell("C24"), cell("C24"))

# ------------------------------------------------------------------- refusals
print("\nBad inputs speak instead of showing an error code")
for case, values, expect in [
    ("noyears", {"B4": 1000000, "B5": 0, "B6": None, "B7": 0, "B8": 10, "B9": 0}, "how many years"),
    ("notarget", {"B4": None, "B5": 400000, "B6": 5, "B7": 0, "B8": 10, "B9": 0}, "aiming for"),
    ("fantasy", {"B4": 1000000, "B5": 0, "B6": 5, "B7": 1000, "B8": 90, "B9": 0}, "50%"),
    ("toolong", {"B4": 1000000, "B5": 0, "B6": 60, "B7": 1000, "B8": 10, "B9": 0}, "40 years"),
]:
    cell, number = run(case, values)
    line = cell("B12")
    report(f"{case}: the line explains it", expect in line, line)
    report(f"{case}: no error code reaches the reader",
           "#" not in cell("B11") and "#" not in line, cell("B11"))

print()
print("FAILURES:", failures if failures else "none")
sys.exit(1 if failures else 0)
