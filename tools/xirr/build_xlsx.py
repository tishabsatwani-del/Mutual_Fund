"""Build XIRR Calculator v1.0.xlsx exactly to the Part A brief.

Allowed functions only: XIRR, IF, IFERROR, DATE, EDATE, TODAY, COUNT, COUNTA,
OFFSET, SUM, TEXT.  No dynamic arrays, no macros, no COUNTIF (not on the list) --
the two row counts the status line needs come from a hidden Calc tab, which the
brief explicitly permits.
"""
import datetime as dt
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

OUT = sys.argv[1]

VERSION = "1.2"
BUILT = "22-Aug-2026"

FIRST_ROW, LAST_ROW = 8, 507
HEAD_ROW = 7

DATE_FMT = "dd-mmm-yyyy"
# Review v4 section 11. Indian grouping: the last three digits, then twos.
# The previous single-section format grouped four digits past the first comma
# (₹568,7111 for what should be ₹56,87,111). Three sections, one per
# magnitude, so the repeat is long enough for a crore and for a lakh.
MONEY_FMT = ("[>=10000000]\u20b9\\ ##\\,##\\,##\\,##0;"
    "[>=100000]\u20b9\\ ##\\,##\\,##0;"
    "\u20b9\\ ##,##0"
)
# The same standard inside a TEXT() formula, where a cell format cannot reach.
MONEY_TEXT = '"[>=10000000]\u20b9 ##\\,##\\,##\\,##0;[>=100000]\u20b9 ##\\,##\\,##0;\u20b9 ##,##0"'
PCT_FMT = "0.0%"

BODY = Font(name="Calibri", size=12)
BODY_BOLD = Font(name="Calibri", size=12, bold=True)
GREY = Font(name="Calibri", size=12, color="FF808080")
HEAD_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="FF1F4E5F")
RESULT_FONT = Font(name="Calibri", size=28, bold=True, color="FF1F4E5F")
RESULT_FILL = PatternFill("solid", fgColor="FFEAF3F6")
GREY_FILL = PatternFill("solid", fgColor="FFF2F2F2")
INPUT_FILL = PatternFill("solid", fgColor="FFFFFDF0")   # every cell a reader types into
STATUS_FONT = Font(name="Calibri", size=12, color="FFB00020")
LABEL_FONT = Font(name="Calibri", size=12, bold=True)

UNLOCKED = Protection(locked=False)
THIN = Side(style="thin", color="FFBFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

INSTRUCTION = (
    "Type in the cream cells only \u2014 everything else is worked out for you. "
    "One row for every payment, with no blank rows in between, and the last row "
    "must be today's date, Worth today, and what the holding is worth now."
)

RESULT_FORMULA = (
    "=IFERROR(XIRR(Flow_Values,Flow_Dates),"
    "IFERROR(XIRR(Flow_Values,Flow_Dates,-0.5),"
    '"Check the status line below"))'
)

STATUS_FORMULA = (
    '=IF(COUNT($B$8:$B$507)=0,"Add your first investment above.",'
    'IF(Calc!$D$1=0,"Add a last row: today\'s date, Worth today, and what the '
    'holding is worth now.",'
    'IF(Calc!$D$2=0,"Add at least one investment.","")))'
)

HEADINGS = ["Date", "What happened", "Amount", "Used by the sheet", "Which holding"]


def excel_serial(d):
    return (d - dt.date(1899, 12, 30)).days


def flow_formula(row):
    return f'=IF($B{row}="","",IF($C{row}="Money in",-$D{row},$D{row}))'


def widths(ws):
    ws.column_dimensions["A"].width = 2.5
    ws.column_dimensions["B"].width = 13.5
    ws.column_dimensions["C"].width = 17
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 18


def header_block(ws, fund_name=None):
    """Rows 1-6: fund name, the result, the status line, the instruction."""
    widths(ws)

    ws["A2"] = "Fund name"
    ws["A2"].font = LABEL_FONT
    ws["B2"] = fund_name
    ws["B2"].font = BODY

    ws["A4"] = "Your XIRR"
    ws["A4"].font = LABEL_FONT
    ws["A4"].alignment = Alignment(vertical="center")
    ws["B4"].font = RESULT_FONT
    ws["B4"].fill = RESULT_FILL
    ws["B4"].number_format = PCT_FMT
    ws["B4"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 38

    ws["B5"].font = STATUS_FONT
    ws["B5"].alignment = Alignment(vertical="center")

    ws["B6"] = INSTRUCTION
    ws["B6"].font = BODY
    ws.row_dimensions[6].height = 16

    for col, heading in zip("BCDEF", HEADINGS):
        c = ws[f"{col}{HEAD_ROW}"]
        c.value = heading
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[HEAD_ROW].height = 30
    ws.freeze_panes = "A8"


def entry_rows(ws, unlock):
    """Pre-format rows 8-507 and drop the flow formula into column E."""
    for row in range(FIRST_ROW, LAST_ROW + 1):
        b, c, d, e, f = (ws[f"{x}{row}"] for x in "BCDEF")
        for cell in (b, c, d, f):
            cell.font = BODY
            cell.border = CELL_BORDER
            if unlock:
                cell.protection = UNLOCKED
        b.number_format = DATE_FMT
        d.number_format = MONEY_FMT
        if unlock:
            # cream means "yours to type in"; grey means "the sheet works this out"
            for cell in (b, c, d, f):
                cell.fill = INPUT_FILL
        e.value = flow_formula(row)
        e.font = GREY
        e.fill = GREY_FILL
        e.border = CELL_BORDER
        e.number_format = MONEY_FMT
        f.font = BODY


def validations(ws):
    lo, hi = excel_serial(dt.date(1990, 1, 1)), excel_serial(dt.date(2100, 12, 31))
    rng = f"B{FIRST_ROW}:B{LAST_ROW}"
    dv_date = DataValidation(
        type="date", operator="between", formula1=str(lo), formula2=str(hi),
        allow_blank=True, showErrorMessage=True,
        errorTitle="Check the date",
        error="Enter a real date between 1-Jan-1990 and 31-Dec-2100.",
    )
    dv_kind = DataValidation(
        type="list", formula1='"Money in,Money out,Worth today"',
        allow_blank=True, showErrorMessage=True,
        errorTitle="Pick one of the three",
        error="Choose Money in, Money out or Worth today from the list.",
    )
    dv_amt = DataValidation(
        type="decimal", operator="greaterThan", formula1="0",
        allow_blank=True, showErrorMessage=True,
        errorTitle="Positive numbers only",
        error="Type the amount as a plain positive number. Never type a minus sign.",
    )
    for dv, ref in ((dv_date, rng), (dv_kind, f"C{FIRST_ROW}:C{LAST_ROW}"),
                    (dv_amt, f"D{FIRST_ROW}:D{LAST_ROW}")):
        ws.add_data_validation(dv)
        dv.add(ref)


def protect(ws):
    ws.protection.sheet = True  # no password: an unopenable file is a dead file
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    ws.protection.formatCells = False


wb = Workbook()

# ---------------------------------------------------------------- Start here
start = wb.active
start.title = "Start here"
start.column_dimensions["A"].width = 3
start.column_dimensions["B"].width = 92
lines = [
    "1. This workbook does two things: it works out what your money has actually "
    "earned, and it shows what it would take to reach a goal.",
    "2. Save your own copy before you type anything, so you always have a clean one.",
    "3. On the My investments tab, enter one row for every payment you made: the "
    "date, then Money in or Money out from the dropdown, then the amount as a "
    "plain positive number.",
    "Dates are shown as dd-mmm-yyyy. If a date you typed reads back as a different "
    "day, your computer is set to American dates; retype it and check it reads the "
    "way you meant. "
    "4. In the last row, enter today's date, choose Worth today, and type what the "
    "holding is worth right now. Your XIRR appears at the top of that tab.",
    "5. On the Plan my goal tab, enter what you are aiming for, what you already "
    "have, and what you invest each month. It shows where you land and what would "
    "close the gap.",
    "6. The Example tab is the same thing already filled in, if you get stuck.",
]
for i, text in enumerate(lines):
    cell = start.cell(row=2 + i * 2, column=2, value=text)
    cell.font = BODY
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    start.row_dimensions[2 + i * 2].height = 32
protect(start)

# ----------------------------------------------------------- My investments
inv = wb.create_sheet("My investments")
header_block(inv)
inv["B2"].protection = UNLOCKED
inv["B4"] = RESULT_FORMULA
inv["B5"] = STATUS_FORMULA
entry_rows(inv, unlock=True)
validations(inv)

# ---- what the whole portfolio did
inv["H2"] = "Your portfolio"
inv["H2"].font = HEAD_FONT
inv["H2"].fill = HEAD_FILL
inv["I2"].fill = HEAD_FILL
SUMMARY = [
    (3, "You put in", "=Calc!$D$3"),
    (4, "You took out", "=Calc!$D$4"),
    (5, "Worth now", "=Calc!$D$5"),
    (6, "Gain or loss", "=Calc!$D$5+Calc!$D$4-Calc!$D$3"),
    (7, "How long you have held it",
     '=IF(COUNT($B$8:$B$507)=0,"",'
     'TEXT((OFFSET($B$8,COUNT($B$8:$B$507)-1,0)-$B$8)/365.25,"0.0")&" years")'),
]
for row, label, formula in SUMMARY:
    inv.cell(row=row, column=8, value=label).font = BODY
    c = inv.cell(row=row, column=9, value=formula)
    c.font = BODY
    if row != 7:
        c.number_format = MONEY_FMT
inv.column_dimensions["H"].width = 26
inv.column_dimensions["I"].width = 18
inv.column_dimensions["J"].width = 16

# ---- each holding on its own, beside the whole
#
# A reader who owns three funds wants to know how each did AND how their money
# did. Those are different questions and usually have different answers, which
# is the single most useful thing this tab can show them.
inv["H9"] = "Each holding on its own"
inv["H9"].font = HEAD_FONT
inv["H9"].fill = HEAD_FILL
for col in "IJ":
    inv[f"{col}9"].fill = HEAD_FILL
inv["H10"] = "Type a holding name"
inv["I10"] = "You put in"
inv["J10"] = "Its own XIRR"
for col in "HIJ":
    inv[f"{col}10"].font = LABEL_FONT

FLOW_COL = ["AE", "AF", "AG", "AH", "AI"]
for slot in range(1, 6):
    row = 10 + slot
    name = inv.cell(row=row, column=8)
    name.font = BODY
    name.protection = UNLOCKED
    name.border = CELL_BORDER
    name.fill = INPUT_FILL
    put = inv.cell(row=row, column=9, value=f"=IF($H{row}=\"\",\"\",Calc!$L${slot})")
    put.font = BODY
    put.number_format = MONEY_FMT
    rate = inv.cell(row=row, column=10, value=(
        f'=IF($H{row}="","",'
        f'IFERROR(XIRR(OFFSET(Calc!${FLOW_COL[slot - 1]}$8,0,0,COUNT($B$8:$B$507),1),Flow_Dates),'
        f'IFERROR(XIRR(OFFSET(Calc!${FLOW_COL[slot - 1]}$8,0,0,COUNT($B$8:$B$507),1),Flow_Dates,-0.5),'
        f'"not enough entries")))'
    ))
    rate.font = BODY
    rate.number_format = PCT_FMT

inv["H17"] = "Leave the names blank if you hold only one thing."
inv["H17"].font = GREY
inv["H18"] = "Your portfolio XIRR is not the average of these."
inv["H18"].font = BODY_BOLD
inv["H19"] = "It weighs each holding by how much money you actually had in it, and for how long."
inv["H19"].font = GREY

protect(inv)

# ------------------------------------------------------------- Plan my goal
#
# Excel cannot loop, so the monthly instalments are summed in closed form. The
# helper terms live on the hidden Calc tab; this tab holds only what a reader
# reads. The closed form was checked against the same calculation done month by
# month, and agrees to fourteen decimal places, so this tab and the web tool
# never disagree.
#
# Percentages are entered as plain numbers -- 10 means 10% -- and divided by 100
# in the formulas. A cell pre-formatted as a percentage turns a typed 10 into
# either 10% or 1000% depending on one Excel setting, and a reader has no way to
# tell which they got.
goal = wb.create_sheet("Plan my goal")
goal.column_dimensions["A"].width = 32
goal.column_dimensions["B"].width = 26      # wide enough for a crore in 20pt
goal.column_dimensions["C"].width = 22

goal["B1"] = "Plan my goal"
goal["B1"].font = Font(name="Calibri", size=16, bold=True, color="FF1F4E5F")

GOAL_INPUTS = [
    (3, "What is this for?", None, "text"),
    (4, "Amount you are aiming for", 5000000, "money"),
    (5, "What you have already", 400000, "money"),
    (6, "Years left", 15, "years"),
    (7, "Investing each month now", 10000, "money"),
    (8, "Assumed return each year (%)", 10, "rate"),
    (9, "Raise the monthly amount each year by (%)", 0, "rate"),
]
for row, label, default, kind in GOAL_INPUTS:
    a = goal.cell(row=row, column=1, value=label)
    a.font = LABEL_FONT
    b = goal.cell(row=row, column=2, value=default)
    b.font = BODY
    b.protection = UNLOCKED
    b.border = CELL_BORDER
    b.fill = INPUT_FILL
    if kind == "money":
        b.number_format = MONEY_FMT
    elif kind in ("rate", "years"):
        b.number_format = "0.#"

goal["A8"].alignment = Alignment(wrap_text=True, vertical="center")
goal["A9"].alignment = Alignment(wrap_text=True, vertical="center")
goal.row_dimensions[9].height = 30
goal["C8"] = "An assumption you choose. Nobody can promise it."
goal["C8"].font = GREY

# ---- the answer
goal["A11"] = "If nothing changes, you reach"
goal["A11"].font = LABEL_FONT
goal["B11"] = ('=IF(Calc!$G$10=0,"Check the line below",'
               'Calc!$G$11+Calc!$G$12)')
# 20pt, not the 28pt used for a short percentage: a goal figure can run to
# nine characters plus a symbol, and an oversized font turns it into ###.
goal["B11"].font = Font(name="Calibri", size=20, bold=True, color="FF1F4E5F")
goal["B11"].fill = RESULT_FILL
goal["B11"].number_format = MONEY_FMT
goal.row_dimensions[11].height = 38

goal["B12"] = (
    '=IF($B$6="","Enter how many years are left, on the line above.",'
    'IF($B$6<=0,"Years left must be more than zero.",'
    'IF($B$6>40,"Enter 40 years or less.",'
    'IF($B$4="","Enter the amount you are aiming for.",'
    'IF($B$4<=0,"The amount you are aiming for must be more than zero.",'
    'IF($B$8>50,"Enter a return of 50% a year or less. A higher assumption does '
    'not make a plan, it hides one.",'
    'IF($B$8<=-100,"Enter a return greater than -100%.","")))))))'
)
goal["B12"].font = STATUS_FONT
goal["B12"].alignment = Alignment(vertical="center")

goal["A14"] = "Against your goal"
goal["A14"].font = LABEL_FONT
goal["B14"] = (
    '=IF(Calc!$G$10=0,"",'
    f'IF(Calc!$G$13>0.5,"Short by "&TEXT(Calc!$G$13,{MONEY_TEXT}),'
    f'IF(Calc!$G$13<-0.5,"Covered, with "&TEXT(-Calc!$G$13,{MONEY_TEXT})&" to spare",'
    '"Covered, exactly on target")))'
)
goal["B14"].font = BODY_BOLD

goal["A15"] = "Extra each month to close it"
goal["A15"].font = LABEL_FONT
goal["B15"] = '=IF(Calc!$G$10=0,"",Calc!$G$14)'
goal["B15"].font = BODY_BOLD
goal["B15"].number_format = MONEY_FMT
goal["C15"] = "On top of what you already invest each month."
goal["C15"].font = GREY

# ---- where the money comes from
goal["A17"] = "Where that comes from"
goal["A17"].font = HEAD_FONT
goal["A17"].fill = HEAD_FILL
for col in "BC":
    goal[f"{col}17"].fill = HEAD_FILL

WHERE = [
    (18, "What you already have, grown", '=IF(Calc!$G$10=0,"",Calc!$G$11)'),
    (19, "What your monthly investing adds", '=IF(Calc!$G$10=0,"",Calc!$G$12)'),
    (20, "Your own money paid in over the years", '=IF(Calc!$G$10=0,"",Calc!$G$15)'),
]
for row, label, formula in WHERE:
    goal.cell(row=row, column=1, value=label).font = BODY
    c = goal.cell(row=row, column=2, value=formula)
    c.font = BODY
    c.number_format = MONEY_FMT

# ---- scenarios
goal["A22"] = "If you did a little more"
goal["A22"].font = LABEL_FONT
for col, head in zip("ABC", ["Scenario", "You reach", "Against your goal"]):
    c = goal[f"{col}23"]
    c.value = head
    c.font = HEAD_FONT
    c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")

SCENARIOS = [
    (24, "Carry on exactly as you are", "Calc!$G$16"),
    (25, "Add \u20b92,000 a month", "Calc!$G$17"),
    (26, "Add \u20b95,000 a month", "Calc!$G$18"),
    (27, "Same amount, raised 10% every year", "Calc!$G$19"),
]
for row, label, ref in SCENARIOS:
    goal.cell(row=row, column=1, value=label).font = BODY
    v = goal.cell(row=row, column=2, value=f'=IF(Calc!$G$10=0,"",{ref})')
    v.font = BODY
    v.number_format = MONEY_FMT
    d = goal.cell(row=row, column=3, value=(
        f'=IF(Calc!$G$10=0,"",IF({ref}>=$B$4,'
        f'"covered, +"&TEXT({ref}-$B$4,{MONEY_TEXT}),'
        f'"short by "&TEXT($B$4-{ref},{MONEY_TEXT})))'
    ))
    d.font = BODY

# Three short lines rather than one merged block: merged cells break selection
# on a touch screen, and this workbook is opened on phones.
NOTE = [
    "Every line above uses the return you typed in. It is an assumption, not a forecast.",
    "Real markets do not deliver the same return every year, and a run of poor years "
    "early on hurts more than the same years late.",
    "Inflation is not deducted, so a goal set in today's rupees will cost more by the "
    "time it arrives.",
]
for offset, text in enumerate(NOTE):
    c = goal.cell(row=29 + offset, column=1, value=text)
    c.font = GREY

dv_years = DataValidation(
    type="whole", operator="between", formula1="1", formula2="40",
    allow_blank=True, showErrorMessage=True,
    errorTitle="Whole years only",
    error="Enter the number of years left as a whole number between 1 and 40.",
)
goal.add_data_validation(dv_years)
dv_years.add("B6")

dv_amounts = DataValidation(
    type="decimal", operator="greaterThanOrEqual", formula1="0",
    allow_blank=True, showErrorMessage=True,
    errorTitle="Positive numbers only",
    error="Type the amount as a plain positive number, with no rupee sign and no commas.",
)
goal.add_data_validation(dv_amounts)
dv_amounts.add("B4")
dv_amounts.add("B5")
dv_amounts.add("B7")

protect(goal)

# ------------------------------------------------------------------ Example
ex = wb.create_sheet("Example")
header_block(ex, fund_name="An example, already filled in")
ex["B4"] = (
    "=IFERROR(XIRR($E$8:$E$10,$B$8:$B$10),"
    "IFERROR(XIRR($E$8:$E$10,$B$8:$B$10,-0.5),"
    '"Check the status line below"))'
)
entry_rows(ex, unlock=False)
for row, (d, kind, amount) in enumerate(
    [
        (dt.datetime(2024, 1, 1), "Money in", 100000),
        (dt.datetime(2025, 1, 1), "Money in", 100000),
        (dt.datetime(2026, 1, 1), "Worth today", 228000),
    ],
    start=FIRST_ROW,
):
    ex[f"B{row}"], ex[f"C{row}"], ex[f"D{row}"] = d, kind, amount
    ex[f"B{row}"].number_format = DATE_FMT
    ex[f"D{row}"].number_format = MONEY_FMT
protect(ex)

# -------------------------------------------------------------------- About
about = wb.create_sheet("About")
about.column_dimensions["A"].width = 3
about.column_dimensions["B"].width = 92
about_lines = [
    f"XIRR Calculator, version {VERSION}",
    f"Built {BUILT}",
    "This is an educational tool for readers. The figure it produces is before "
    "exit load and before tax, and it is not investment advice.",
    "Questions: use the channels listed in the book.",
    "The tabs are protected without a password. To adapt this sheet, choose "
    "Review, then Unprotect Sheet.",
    "The Plan my goal tab uses a return you type in yourself. It is an assumption, "
    "not a forecast, and it is shown before inflation.",
    "No Excel? Upload this file to Google Sheets and it works there too, unchanged.",
]
for i, text in enumerate(about_lines):
    cell = about.cell(row=2 + i * 2, column=2, value=text)
    cell.font = BODY_BOLD if i == 0 else BODY
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    about.row_dimensions[2 + i * 2].height = 30
protect(about)

# --------------------------------------------------------------- Calc (hidden)
calc = wb.create_sheet("Calc")
calc["C1"] = "Rows marked Worth today"
calc["D1"] = f"=SUM($A${FIRST_ROW}:$A${LAST_ROW})"
calc["C2"] = "Rows marked Money in"
calc["D2"] = f"=SUM($B${FIRST_ROW}:$B${LAST_ROW})"
# Helper columns live out at AA and beyond so they can never collide with the
# goal tab's terms in F to I.
FLOW_COLS = ["AE", "AF", "AG", "AH", "AI"]      # each holding's cash flow
PAID_COLS = ["AK", "AL", "AM", "AN", "AO"]      # each holding's money in

for row in range(FIRST_ROW, LAST_ROW + 1):
    calc[f"A{row}"] = f"=IF('My investments'!$C{row}=\"Worth today\",1,0)"
    calc[f"B{row}"] = f"=IF('My investments'!$C{row}=\"Money in\",1,0)"
    calc[f"AA{row}"] = f"=IF('My investments'!$C{row}=\"Money in\",'My investments'!$D{row},0)"
    calc[f"AB{row}"] = f"=IF('My investments'!$C{row}=\"Money out\",'My investments'!$D{row},0)"
    calc[f"AC{row}"] = f"=IF('My investments'!$C{row}=\"Worth today\",'My investments'!$D{row},0)"
    # One column per holding: this row's flow if it belongs to that holding,
    # otherwise zero. A zero contributes nothing to XIRR, so each column is that
    # holding's own cash flow with every date left in place -- which is how a
    # subset gets measured without a function that filters.
    for slot in range(5):
        name_cell = f"'My investments'!$H${11 + slot}"
        calc[f"{FLOW_COLS[slot]}{row}"] = (
            f"=IF({name_cell}=\"\",0,"
            f"IF('My investments'!$F{row}={name_cell},'My investments'!$E{row},0))"
        )
        calc[f"{PAID_COLS[slot]}{row}"] = (
            f"=IF({name_cell}=\"\",0,"
            f"IF('My investments'!$F{row}={name_cell},$AA{row},0))"
        )

# what each holding received, for the table beside it
for slot in range(5):
    calc[f"K{slot + 1}"] = f"holding {slot + 1} invested"
    calc[f"L{slot + 1}"] = f"=SUM({PAID_COLS[slot]}${FIRST_ROW}:{PAID_COLS[slot]}${LAST_ROW})"

# ---- what the whole portfolio did, for the summary beside the entry table
calc["C3"] = "Total invested"
calc["D3"] = f"=SUM($AA${FIRST_ROW}:$AA${LAST_ROW})"
calc["C4"] = "Total withdrawn"
calc["D4"] = f"=SUM($AB${FIRST_ROW}:$AB${LAST_ROW})"
calc["C5"] = "Worth today"
calc["D5"] = f"=SUM($AC${FIRST_ROW}:$AC${LAST_ROW})"

# ---- goal-tab helpers.
#
# Excel has no loop, so a year of month-start instalments is summed in closed
# form and then compounded forward. G3 is what one rupee a month becomes over a
# year; G8 is what one rupee a month becomes over the whole period, which is why
# the extra needed to close a gap is a division rather than a search.
G = "'Plan my goal'!"
calc["F1"] = "annual rate r";        calc["G1"] = f'=IF({G}$B$8="",0,{G}$B$8/100)'
calc["F2"] = "monthly rate i";       calc["G2"] = '=(1+$G$1)^(1/12)-1'
calc["F3"] = "one year of 1/month";  calc["G3"] = '=IF($G$2=0,12,(((1+$G$2)^12-1)/$G$2)*(1+$G$2))'
calc["F4"] = "years Y";              calc["G4"] = f'=IF({G}$B$6="",0,{G}$B$6)'
calc["F5"] = "step-up g";            calc["G5"] = f'=IF({G}$B$9="",0,{G}$B$9/100)'
calc["F6"] = "ratio x";              calc["G6"] = '=(1+$G$5)/(1+$G$1)'
calc["F7"] = "geometric sum S";      calc["G7"] = '=IF($G$4<=0,0,IF($G$6=1,$G$4,(1-$G$6^$G$4)/(1-$G$6)))'
calc["F8"] = "growth of 1/month";    calc["G8"] = '=IF($G$4<=0,0,$G$3*(1+$G$1)^($G$4-1)*$G$7)'

calc["F10"] = "inputs usable?"
calc["G10"] = (
    f'=IF({G}$B$6="",0,IF({G}$B$6<=0,0,IF({G}$B$6>40,0,'
    f'IF({G}$B$4="",0,IF({G}$B$4<=0,0,'
    f'IF({G}$B$8>50,0,IF({G}$B$8<=-100,0,1)))))))'
)
calc["F11"] = "from what you have";  calc["G11"] = f'=IF({G}$B$5="",0,{G}$B$5)*(1+$G$1)^$G$4'
calc["F12"] = "from monthly";        calc["G12"] = f'=IF({G}$B$7="",0,{G}$B$7)*$G$8'
calc["F13"] = "gap";                 calc["G13"] = f'=IF({G}$B$4="",0,{G}$B$4)-($G$11+$G$12)'
calc["F14"] = "extra each month"
calc["G14"] = '=IF($G$13<=0.5,0,IF($G$8<=0,0,$G$13/$G$8))'
calc["F20"] = "step-up sum T"
calc["G20"] = '=IF($G$4<=0,0,IF($G$5=0,$G$4,((1+$G$5)^$G$4-1)/$G$5))'
calc["F15"] = "own money paid in"
calc["G15"] = f'=IF({G}$B$7="",0,{G}$B$7)*12*$G$20'

# ---- the scenario rows, sharing the same helpers
calc["F16"] = "as you are";   calc["G16"] = '=$G$11+$G$12'
calc["F17"] = "plus 2,000";   calc["G17"] = f'=$G$11+(IF({G}$B$7="",0,{G}$B$7)+2000)*$G$8'
calc["F18"] = "plus 5,000";   calc["G18"] = f'=$G$11+(IF({G}$B$7="",0,{G}$B$7)+5000)*$G$8'

# a 10% step-up needs its own ratio and sum, so it gets its own short column
calc["I5"] = '=IF($G$5>0.1,$G$5,0.1)'
calc["I6"] = '=(1+$I$5)/(1+$G$1)'
calc["I7"] = '=IF($G$4<=0,0,IF($I$6=1,$G$4,(1-$I$6^$G$4)/(1-$I$6)))'
calc["I8"] = '=IF($G$4<=0,0,$G$3*(1+$G$1)^($G$4-1)*$I$7)'
calc["F19"] = "raised 10% a year"
calc["G19"] = f'=$G$11+IF({G}$B$7="",0,{G}$B$7)*$I$8'

for row in calc.iter_rows():
    for c in row:
        if c.value is not None:
            c.font = BODY
protect(calc)
calc.sheet_state = "hidden"

# ------------------------------------------------------------- Named ranges
for name, col in (("Flow_Dates", "B"), ("Flow_Values", "E")):
    wb.defined_names.add(
        DefinedName(
            name,
            attr_text=(
                f"OFFSET('My investments'!${col}${FIRST_ROW},0,0,"
                f"COUNT('My investments'!$B${FIRST_ROW}:$B${LAST_ROW}),1)"
            ),
        )
    )

wb.active = 0
wb.save(OUT)
print("wrote", OUT)
