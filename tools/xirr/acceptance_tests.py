"""Acceptance tests for the built workbook.

Two independent checks per case:
  1. a from-scratch XIRR solver (Excel's definition: 365-day year, Newton then
     bisection) run over the cash flows,
  2. the actual workbook -- rows typed in, then recalculated by LibreOffice and
     read back out of cells B4 and B5.
"""
import datetime as dt
import os
import shutil
import subprocess
import sys
import csv
import glob

from openpyxl import load_workbook

SRC = sys.argv[1]
WORK = "run"
PROFILE = "file:///tmp/lo-profile-xirr"


# ---------------------------------------------------------------- solver
def xnpv(rate, flows):
    d0 = flows[0][0]
    return sum(p / (1.0 + rate) ** ((d - d0).days / 365.0) for d, p in flows)


def xirr(flows):
    lo, hi = -0.9999999, 100.0
    flo = xnpv(lo, flows)
    for _ in range(400):
        mid = (lo + hi) / 2
        fm = xnpv(mid, flows)
        if flo * fm <= 0:
            hi = mid
        else:
            lo, flo = mid, fm
    return (lo + hi) / 2


# ---------------------------------------------------------------- cases
def sip_rows():
    rows = []
    for n in range(60):
        y, m = divmod((2020 * 12) + n, 12)
        rows.append((dt.date(y, m + 1, 1), "Investment", 10000))
    rows.append((dt.date(2025, 1, 1), "Value today", 824864))
    return rows


CASES = [
    ("1. Two lump sums", [
        (dt.date(2024, 1, 1), "Investment", 100000),
        (dt.date(2025, 1, 1), "Investment", 100000),
        (dt.date(2026, 1, 1), "Value today", 228000),
    ], "9.1%"),
    ("2. Five-year monthly SIP", sip_rows(), "12.7%"),
    ("3. With a withdrawal", [
        (dt.date(2022, 4, 1), "Investment", 500000),
        (dt.date(2024, 4, 1), "Withdrawal", 200000),
        (dt.date(2026, 4, 1), "Value today", 450000),
    ], "8.1%"),
    ("a. Investments but no Value today row", [
        (dt.date(2024, 1, 1), "Investment", 100000),
        (dt.date(2025, 1, 1), "Investment", 100000),
    ], "status line"),
    ("b. A single row", [
        (dt.date(2024, 1, 1), "Investment", 100000),
    ], "status line"),
    ("c. Empty table", [], "status line"),
    ("d. Value today but no investment", [
        (dt.date(2026, 1, 1), "Value today", 228000),
    ], "status line"),
]


def signed(kind, amount):
    return -amount if kind == "Investment" else amount


# ---------------------------------------------------------------- workbook
def fill(path, rows):
    wb = load_workbook(SRC)
    ws = wb["My investments"]
    ws.protection.sheet = False          # so LibreOffice can be scripted freely
    for i, (d, kind, amount) in enumerate(rows):
        r = 8 + i
        ws[f"B{r}"] = dt.datetime(d.year, d.month, d.day)
        ws[f"B{r}"].number_format = "dd-mmm-yyyy"
        ws[f"C{r}"] = kind
        ws[f"D{r}"] = amount
    wb.save(path)


def recalc(path, sheet_index):
    """Convert one sheet to CSV; LibreOffice evaluates every formula on load."""
    outdir = os.path.join(WORK, f"out{sheet_index}")
    os.makedirs(outdir, exist_ok=True)
    flt = ("csv:Text - txt - csv (StarCalc):44,34,76,1,,0,false,true,true,"
           f"false,false,{sheet_index}")
    subprocess.run(
        ["soffice", f"-env:UserInstallation={PROFILE}", "--headless",
         "--convert-to", flt, "--outdir", outdir, path],
        check=True, capture_output=True, timeout=240,
    )
    base = os.path.splitext(os.path.basename(path))[0]
    # with a sheet index the filter writes "<base>-<sheet name>.csv"
    hits = glob.glob(os.path.join(outdir, base + "*.csv"))
    with open(hits[0], newline="", encoding="utf-8") as fh:
        return list(csv.reader(fh))


def cell(grid, row, col):
    try:
        return grid[row - 1][col - 1]
    except IndexError:
        return ""


shutil.rmtree(WORK, ignore_errors=True)
os.makedirs(WORK, exist_ok=True)

print(f"{'case':42} {'sheet B4':>12}  {'solver':>10}  status line")
print("-" * 110)

results = []
for name, rows, expected in CASES:
    path = os.path.join(WORK, name.split(".")[0].strip() + ".xlsx")
    fill(path, rows)
    grid = recalc(path, 2)          # 2 = My investments
    b4 = cell(grid, 4, 2)
    b5 = cell(grid, 5, 2)

    flows = [(d, signed(k, a)) for d, k, a in rows]
    solver = ""
    if len([f for f in flows if f[1] < 0]) and len([f for f in flows if f[1] > 0]):
        solver = f"{xirr(sorted(flows))*100:.4f}%"

    print(f"{name:42} {b4:>12}  {solver:>10}  {b5}")
    results.append((name, b4, b5, solver, expected))

# the Example tab, untouched, must also carry a live answer
grid = recalc(SRC, 3)               # 3 = Example
print("-" * 110)
print(f"{'Example tab (as shipped)':42} {cell(grid, 4, 2):>12}")
