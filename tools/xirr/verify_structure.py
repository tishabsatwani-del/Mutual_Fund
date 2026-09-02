"""Assert the shipped file matches the brief, clause by clause."""
import re
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook

PATH = sys.argv[1]
wb = load_workbook(PATH)
fails = []


def check(label, ok, detail=""):
    print(f"{'PASS' if ok else 'FAIL'}  {label}{('  -- ' + detail) if detail else ''}")
    if not ok:
        fails.append(label)


# tab order and visibility
check("five visible tabs, in order",
      [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
      == ["Start here", "My investments", "Plan my goal", "Example", "About"])
check("only Calc is hidden",
      [ws.title for ws in wb.worksheets if ws.sheet_state != "visible"] == ["Calc"])

inv = wb["My investments"]

# named ranges
names = {n: d.value for n, d in wb.defined_names.items()}
for n, col in (("Flow_Dates", "B"), ("Flow_Values", "E")):
    check(f"named range {n}", names.get(n) ==
          f"OFFSET('My investments'!${col}$8,0,0,COUNT('My investments'!$B$8:$B$507),1)")

# result and status formulas
check("B4 double XIRR with -0.5 fallback",
      inv["B4"].value.count("XIRR") == 2 and "-0.5" in inv["B4"].value
      and "Check the status line below" in inv["B4"].value)
check("B4 formatted as percent, one decimal", inv["B4"].number_format == "0.0%")
for phrase in ("Add your first investment above.",
               "Add a last row: today's date, Worth today,",
               "Add at least one investment."):
    check(f"status line: {phrase[:34]}...", phrase in inv["B5"].value)

# entry table
check("headings on row 7",
      [inv[f"{c}7"].value for c in "BCDEF"]
      == ["Date", "What happened", "Amount", "Used by the sheet", "Which holding"])
check("500 pre-formatted rows, 8 to 507",
      inv["E8"].value == '=IF($B8="","",IF($C8="Money in",-$D8,$D8))'
      and inv["E507"].value == '=IF($B507="","",IF($C507="Money in",-$D507,$D507))')
check("dates shown dd-mmm-yyyy", inv["B8"].number_format == "dd-mmm-yyyy")
check("freeze panes at row 8", inv.freeze_panes == "A8")
check("no merged cells anywhere",
      all(not ws.merged_cells.ranges for ws in wb.worksheets))
check("no conditional formatting",
      all(not list(ws.conditional_formatting) for ws in wb.worksheets))

# fonts: nothing under 12pt
small = [f"{ws.title}!{c.coordinate}" for ws in wb.worksheets for row in ws.iter_rows()
         for c in row if c.value is not None and c.font and c.font.size
         and c.font.size < 12]
check("no font below 12pt", not small, ", ".join(small[:4]))

# validation
dvs = {dv.type: dv for dv in inv.data_validations.dataValidation}
check("date validation 1990-2100",
      dvs["date"].formula1 == "32874" and dvs["date"].formula2 == "73415")
check("three dropdown options only",
      dvs["list"].formula1 == '"Money in,Money out,Worth today"')
check("amount must be positive",
      dvs["decimal"].operator == "greaterThan" and dvs["decimal"].formula1 == "0")

# protection
check("every sheet protected", all(ws.protection.sheet for ws in wb.worksheets))
check("no sheet password",
      all(not ws.protection.password for ws in wb.worksheets))
unlocked = [c.coordinate for row in inv.iter_rows(min_row=1, max_row=507, max_col=10)
            for c in row if c.protection and c.protection.locked is False]
check("inputs unlocked: fund name, the entry table and the holding names",
      set(unlocked) == ({"B2"} | {f"{c}{r}" for c in "BCDF" for r in range(8, 508)}
                        | {f"H{r}" for r in range(11, 16)}),
      f"{len(unlocked)} cells")
check("result cell and column E locked",
      inv["B4"].protection.locked is not False and inv["E8"].protection.locked is not False)
check("Example tab fully locked",
      not [c for row in wb["Example"].iter_rows() for c in row
           if c.protection and c.protection.locked is False])

check("each holding can be measured on its own",
      "XIRR" in str(inv["J11"].value) and "not enough entries" in str(inv["J11"].value))
check("the portfolio summary is on the working tab",
      [inv[f"H{r}"].value for r in range(3, 8)]
      == ["You put in", "You took out", "Worth now", "Gain or loss", "How long you have held it"])
check("the sheet says the portfolio is not an average of the holdings",
      "not the average" in str(inv["H18"].value))

# the goal tab
goal = wb["Plan my goal"]
calc = wb["Calc"]
check("goal inputs are the only unlocked cells on the goal tab",
      set(c.coordinate for row in goal.iter_rows() for c in row
          if c.protection and c.protection.locked is False)
      == {"B3", "B4", "B5", "B6", "B7", "B8", "B9"})
check("the projected value is locked", goal["B11"].protection.locked is not False)
# The one place a typed input could quietly produce a wrong number: a holding
# name that does not match is silently excluded from that holding's XIRR.
hold_dv = [dv for dv in inv.data_validations.dataValidation
           if dv.type == "list" and "F8" in dv.sqref]
# The SHIPPED file is the hardened one, and the round trip through LibreOffice
# normalises openpyxl's "=$H$11:$H$15" to the stored form "$H$11:$H$15". Assert
# the range, not the punctuation, or this check passes on the build and fails on
# the artifact readers actually download.
check("the holding column offers the names it must match",
      len(hold_dv) == 1 and hold_dv[0].formula1.lstrip("=") == "$H$11:$H$15",
      hold_dv[0].formula1 if hold_dv else "no list validation on F")
check("and warns rather than blocking, so a blank list is still usable",
      bool(hold_dv) and hold_dv[0].errorStyle == "warning",
      hold_dv[0].errorStyle if hold_dv else "")
check("no example holdings are seeded into the reader's own sheet",
      all(inv[f"H{r}"].value in (None, "") for r in range(11, 16)),
      str([inv[f"H{r}"].value for r in range(11, 16)]))

# Review v4 §11's caps, on the sheet as well as on the web. The step-up had
# neither a validation nor a gate, so an absurd raise printed an enormous figure
# with nothing flagged -- the workbook's version of the web planner's 68-digit
# defect.
def goal_dv_for(cell):
    for dv in goal.data_validations.dataValidation:
        if cell in dv.sqref:
            return dv
    return None

for cell, lo, hi in (("B8", "0", "30"), ("B9", "0", "25")):
    dv = goal_dv_for(cell)
    check(f"goal {cell} is bounded {lo} to {hi}",
          dv is not None and dv.formula1 == lo and dv.formula2 == hi,
          f"{cell}: {dv.formula1 if dv else 'no validation'}-{dv.formula2 if dv else ''}")
    check(f"goal {cell} explains itself when refused",
          dv is not None and bool(dv.error) and bool(dv.errorTitle))

# and the gate behind them, which a paste cannot bypass
gate = calc["G10"].value
for ref, bound in (("$B$8", "30"), ("$B$9", "25")):
    check(f"the gate tests {ref} against {bound}",
          f"{ref}>{bound}" in gate.replace("'Plan my goal'!", ""), gate[:120])

check("years are limited to whole numbers",
      any(dv.type == "whole" for dv in goal.data_validations.dataValidation))
check("the goal tab never shows a bare error",
      'Check the line below' in str(goal["B11"].value))
check("the goal tab says the return is an assumption",
      any("assumption, not a forecast" in str(c.value)
          for row in goal.iter_rows() for c in row if c.value))

# ---------------------------------------------------------------- presentation
#
# Every check below is about whether a reader can READ what the sheet says, and
# every one of them has been broken at least once by the recalculation pass
# rather than by the build -- which is why they are asserted on the SHIPPED
# file rather than trusted from build_xlsx.py.

# The header block. 28 points on rows 1-5 so a label and its figure sit on one
# line; row 4 taller still, because it carries the result in a 28-point font and
# a 28-point row cannot hold a 28-point glyph.
for row, want in {1: 28, 2: 28, 3: 28, 5: 28}.items():
    dim = inv.row_dimensions[row]
    check(f"row {row} of My investments is 28 points",
          dim.height is not None and abs(dim.height - want) < 0.5, f"height {dim.height}")
check("row 4 is taller still, so the 28-point result is not clipped",
      inv.row_dimensions[4].height is not None and inv.row_dimensions[4].height >= 36,
      f"height {inv.row_dimensions[4].height}")
check("the header block is aligned to the middle of its own row height",
      all(inv[f"{c}{r}"].alignment.vertical == "center" for c in "AB" for r in (1, 2, 3, 4, 5)))

# A label spills into the cell to its right only while that cell is EMPTY, so a
# label column too narrow for its own text reads correctly on an untouched sheet
# and clips the moment the sheet has anything in it. That is how "Your XIRR"
# shipped as "Yo": correct in every empty-sheet check, wrong for every reader.
LABELS = [inv[f"A{r}"].value for r in (2, 4) if inv[f"A{r}"].value]
longest = max((len(str(v)) for v in LABELS), default=0)
check("the label column is wider than its own longest label",
      inv.column_dimensions["A"].width is not None and
      inv.column_dimensions["A"].width >= longest + 1,
      f"width {inv.column_dimensions['A'].width} for {longest} characters: {LABELS}")

# The instruction runs to about two hundred characters and column B is 13.5 of
# them. Wrapping confined it to that width and the sheet showed its first four
# words; merging would give it room, but this workbook forbids merged cells. It
# spills, which needs the cells beside it to stay empty.
check("the instruction is not wrapped inside a column too narrow for it",
      not inv["B6"].alignment.wrap_text,
      f"wrap_text={inv['B6'].alignment.wrap_text}, column B is {inv.column_dimensions['B'].width}")
check("and nothing blocks it from spilling across the row",
      all(inv[f"{c}6"].value in (None, "") for c in "CDEF"),
      str([inv[f"{c}6"].value for c in "CDEF"]))

# Editable cells are the ONLY yellow ones, and the yellow has to survive a bad
# screen: the previous FFFFFDF0 was indistinguishable from white at an angle,
# which made "type in the coloured cells only" an instruction the sheet did not
# help anyone follow.
check("every cell a reader types into is filled #FEF9C3",
      all(inv[f"{c}8"].fill.fgColor.rgb == "FFFEF9C3" for c in "BCDF") and
      inv["H11"].fill.fgColor.rgb == "FFFEF9C3",
      inv["B8"].fill.fgColor.rgb)
check("and every formula output is locked",
      all(inv[ref].protection.locked for ref in
          ("B4", "B5", "E8", "E507", "I3", "I4", "I5", "I6", "I7", "J11", "J15")))
check("while the cells a reader types into are not",
      not any(inv[ref].protection.locked for ref in ("B2", "B8", "C8", "D8", "F8", "H11")))
check("every tab is protected", all(ws.protection.sheet for ws in wb.worksheets
                                    if ws.title != "Calc"))

# Instruction blocks wrap, and their rows are tall enough to SHOW the wrap. A
# wrapped line in a row too short for it is hidden, not wrapped.
for tab in ("Start here", "About"):
    sheet = wb[tab]
    blocks = [(r, sheet.cell(row=r, column=2)) for r in range(2, 20)
              if isinstance(sheet.cell(row=r, column=2).value, str)]
    check(f"every instruction block on {tab} wraps",
          blocks and all(c.alignment.wrap_text for _, c in blocks), f"{len(blocks)} blocks")
    check(f"and each row on {tab} is tall enough to show what it wraps to",
          all(sheet.row_dimensions[r].height is None or
              sheet.row_dimensions[r].height >= 15.5 * (1 + len(c.value) // 88)
              for r, c in blocks),
          str([(r, sheet.row_dimensions[r].height, len(c.value)) for r, c in blocks]))

# The marketing badge ("100% Standalone & Private ...") was removed after an
# external audit. It must not come back anywhere in the workbook; the About tab
# says the same thing in a plain sentence instead.
BADGE_MARK = "Standalone & Private"
where = [f"{ws.title}!{c.coordinate}" for ws in wb.worksheets
         for r in ws.iter_rows() for c in r
         if isinstance(c.value, str) and BADGE_MARK in c.value]
check("the privacy badge appears nowhere in the workbook", not where, ", ".join(where))
about = wb["About"]
about_text = [c.value for r in about.iter_rows() for c in r if isinstance(c.value, str)]
check("the About tab says plainly that it connects to nothing",
      "It works with no internet connection, and nothing in it connects to anything."
      in about_text)

# ONE version string. The site declares it in tool/app.js and the workbook's
# About tab must carry the same one, so the two can never drift apart.
def site_version(fallback="2.1"):
    app_js = Path(__file__).resolve().parents[2] / "tool" / "app.js"
    try:
        m = re.search(r"var\s+VERSION\s*=\s*['\"]([^'\"]+)['\"]", app_js.read_text(encoding="utf-8"))
    except OSError:
        m = None
    return m.group(1) if m else fallback

SITE_VERSION = site_version()
check(f"the About tab names the site's version ({SITE_VERSION})",
      f"Where You Stand \u2014 XIRR Calculator, version {SITE_VERSION}" in about_text,
      str([t for t in about_text if "version" in t]))

# forbidden functions, anywhere in the workbook
ALLOWED = {"XIRR", "IF", "IFERROR", "DATE", "EDATE", "TODAY", "COUNT", "COUNTA",
           "OFFSET", "SUM", "TEXT"}
used = set()
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                used |= set(re.findall(r"([A-Z][A-Z0-9_.]*)\s*\(", c.value.upper()))
for n, formula in names.items():
    used |= set(re.findall(r"([A-Z][A-Z0-9_.]*)\s*\(", formula.upper()))
check("only allowed functions used", used <= ALLOWED, f"used: {sorted(used)}")
check("no _xlfn future-function prefixes anywhere",
      not any("_xlfn" in str(c.value) for ws in wb.worksheets
              for row in ws.iter_rows() for c in row))

# file-level
with zipfile.ZipFile(PATH) as z:
    parts = z.namelist()
check("no macros in the package",
      not [p for p in parts if "vbaProject" in p or p.endswith(".bin")], "")
check("no external links or data connections",
      not [p for p in parts if "externalLink" in p or "connections" in p])
check("single file, no workbook protection", wb.security is None
      or not getattr(wb.security, "lockStructure", False))

print()
print("FAILURES:", fails if fails else "none")
sys.exit(1 if fails else 0)
