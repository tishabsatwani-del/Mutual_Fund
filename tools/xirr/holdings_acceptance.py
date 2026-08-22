"""Acceptance tests for the per-holding table on the My investments tab.

Each holding is measured by feeding XIRR a column that holds that holding's own
cash flow and a zero everywhere else. Zeros contribute nothing to the present
value, so the answer is the holding's own return -- but that is the kind of trick
that deserves a test rather than trust.

Expected values come from a bisection solver written here, not from the sheet.
"""
import csv
import datetime as dt
import glob
import os
import shutil
import subprocess
import sys

from openpyxl import load_workbook

SRC = sys.argv[1]
WORK = "holdings_run"
PROFILE = "file:///tmp/lo-profile-xirr"
SHEET = 2                      # Start here, My investments, ...

failures = []


def report(name, ok, detail=""):
    print(f"{'PASS' if ok else 'FAIL'}  {name}{('   -- ' + detail) if detail else ''}")
    if not ok:
        failures.append(name)


def close(name, actual, expected, tol):
    ok = actual is not None and abs(actual - expected) <= tol
    report(name, ok, f"sheet {actual}, solver {expected:.4f}" if actual is not None else "no value")


def xirr(flows):
    """Excel's definition, found by bisection. Independent of the workbook."""
    def npv(rate):
        d0 = flows[0][0]
        return sum(p / (1 + rate) ** ((d - d0).days / 365.0) for d, p in flows)
    lo, hi = -0.9999, 10.0
    flo = npv(lo)
    for _ in range(300):
        mid = (lo + hi) / 2
        fm = npv(mid)
        if flo * fm <= 0:
            hi = mid
        else:
            lo, flo = mid, fm
    return (lo + hi) / 2


ROWS = [
    (dt.date(2021, 1, 1), "Investment", 500000, "Fund A"),
    (dt.date(2026, 1, 1), "Value today", 700000, "Fund A"),
    (dt.date(2025, 1, 1), "Investment", 50000, "Fund B"),
    (dt.date(2026, 1, 1), "Value today", 75000, "Fund B"),
]


def signed(kind, amount):
    return -amount if kind == "Investment" else amount


shutil.rmtree(WORK, ignore_errors=True)
os.makedirs(WORK, exist_ok=True)

wb = load_workbook(SRC)
ws = wb["My investments"]
ws.protection.sheet = False
for i, (d, kind, amount, holding) in enumerate(ROWS):
    r = 8 + i
    ws[f"B{r}"] = dt.datetime(d.year, d.month, d.day)
    ws[f"C{r}"] = kind
    ws[f"D{r}"] = amount
    ws[f"F{r}"] = holding
ws["H11"] = "Fund A"
ws["H12"] = "Fund B"
for ref in ["B4", "J11", "J12", "J13"]:
    ws[ref].number_format = "0.0000%"
for ref in ["I3", "I4", "I5", "I6", "I11", "I12"]:
    ws[ref].number_format = "0.00"
path = os.path.join(WORK, "holdings.xlsx")
wb.save(path)

outdir = os.path.join(WORK, "out")
os.makedirs(outdir, exist_ok=True)
flt = ("csv:Text - txt - csv (StarCalc):44,34,76,1,,0,false,true,true,"
       f"false,false,{SHEET}")
subprocess.run(
    ["soffice", f"-env:UserInstallation={PROFILE}", "--headless",
     "--convert-to", flt, "--outdir", outdir, path],
    check=True, capture_output=True, timeout=300,
)
with open(glob.glob(os.path.join(outdir, "*.csv"))[0], newline="", encoding="utf-8") as fh:
    grid = list(csv.reader(fh))


def cell(ref):
    col = 0
    i = 0
    while ref[i].isalpha():
        col = col * 26 + (ord(ref[i].upper()) - 64)
        i += 1
    row = int(ref[i:])
    try:
        return grid[row - 1][col - 1]
    except IndexError:
        return ""


def number(ref):
    raw = cell(ref).replace(",", "").replace("₹", "").replace("%", "").strip()
    try:
        return float(raw)
    except ValueError:
        return None


print("\nEach holding measured on its own")
a = [(d, signed(k, amt)) for d, k, amt, h in ROWS if h == "Fund A"]
b = [(d, signed(k, amt)) for d, k, amt, h in ROWS if h == "Fund B"]
close("Fund A's own XIRR", number("J11"), xirr(a) * 100, 0.02)
close("Fund B's own XIRR", number("J12"), xirr(b) * 100, 0.02)
close("Fund A's money in", number("I11"), 500000, 1)
close("Fund B's money in", number("I12"), 50000, 1)
report("an unused holding row stays blank", cell("J13").strip() == "", repr(cell("J13")))

print("\nThe whole portfolio")
whole = [(d, signed(k, amt)) for d, k, amt, h in ROWS]
combined = xirr(whole) * 100
close("portfolio XIRR is calculated from every flow", number("B4"), combined, 0.02)
report("the portfolio is not the average of its holdings",
       abs(combined - (xirr(a) + xirr(b)) / 2 * 100) > 10,
       f"portfolio {combined:.1f}%, average of the two {((xirr(a) + xirr(b)) / 2 * 100):.1f}%")
close("total invested", number("I3"), 550000, 1)
close("total withdrawn", number("I4"), 0, 1)
close("value today", number("I5"), 775000, 1)
close("gain or loss", number("I6"), 225000, 1)
report("holding period is stated in years", "years" in cell("I7"), cell("I7"))

print()
print("FAILURES:", failures if failures else "none")
sys.exit(1 if failures else 0)
