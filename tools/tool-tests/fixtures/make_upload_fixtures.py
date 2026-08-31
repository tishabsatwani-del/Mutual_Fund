#!/usr/bin/env python3
"""Files for tools/tool-tests/upload.test.js, in the shapes real ones arrive in.

Every one of these is modelled on a file a reader would actually download:
NSE Indices ships thirteen columns with a title above them, AMFI ships
semicolons, a fund house ships an .xlsx whose first three rows are a logo and
a report title, and somebody will always try a PDF. Nothing here is invented
market history -- the numbers grow at a fixed rate so every expected figure in
the test is arithmetic rather than a recording of what the code did.
"""
import datetime as dt, os, sys, zlib

OUT = sys.argv[1] if len(sys.argv) > 1 else '/tmp/prc/upload'
os.makedirs(OUT, exist_ok=True)
M = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

def days(from_y, to_y):
    d, end = dt.date(from_y, 1, 1), dt.date(to_y, 1, 1)
    while d <= end:
        yield d
        d += dt.timedelta(days=1)

def dmy(d):   return '%02d-%s-%d' % (d.day, M[d.month - 1], d.year)
def iso(d):   return d.isoformat()

def write(name, text):
    p = os.path.join(OUT, name)
    with open(p, 'w', newline='\n') as f:
        f.write(text)
    return p

# 1 -- NSE Indices historical export: a title, a blank, then thirteen columns.
#      Nine of them are numeric and only one is the index value.
rows = ['Nifty 50 Total Returns Index', 'Historical Index Data', '',
        'Date,Index Name,Open Index Value,High Index Value,Low Index Value,'
        'Closing Index Value,Points Change,Change(%),Volume,'
        'Turnover (Rs. Cr.),P/E,P/B,Div Yield']
v = 1000.0
for d in days(2010, 2025):
    rows.append('%s,Nifty 50 TRI,%.2f,%.2f,%.2f,%.2f,-4.55,-0.02,238412290,18500.25,22.41,4.13,1.24'
                % (dmy(d), v * 0.999, v * 1.004, v * 0.996, v))
    v *= (1.12) ** (1 / 365.2425)
write('nse-nifty50-tri.csv', '\n'.join(rows))

# 2 -- AMFI, semicolons, one scheme, NAV among seven columns
rows = ['Scheme Code;Scheme Name;ISIN Div Payout/ISIN Growth;ISIN Div Reinvestment;'
        'Net Asset Value;Repurchase Price;Sale Price;Date']
v = 100.0
for d in days(2010, 2025):
    rows.append('120503;Alpha Bluechip Fund - Direct Plan - Growth;INF204K01XI3;-;'
                '%.4f;%.4f;%.4f;%s' % (v, v * 0.999, v * 1.001, dmy(d)))
    v *= (1.14) ** (1 / 365.2425)
write('amfi-alpha-nav.csv', '\n'.join(rows))

# 3 -- a Zerodha tradebook. Must be refused, and named as one.
rows = ['symbol,isin,trade_date,exchange,segment,series,trade_type,auction,'
        'quantity,price,trade_id,order_id,order_execution_time']
for i, d in enumerate(days(2023, 2024)):
    rows.append('INFY,INE009A01021,%s,NSE,EQ,EQ,%s,false,%d,%.2f,100%d,230%d,%sT09:30:00'
                % (iso(d), 'sell' if i % 2 else 'buy', 10 + i % 40, 1400 + i % 90, i, i, iso(d)))
write('zerodha-tradebook.csv', '\n'.join(rows))

# 4 -- right headings, wrong contents
rows = ['Date,NAV']
for d in list(days(2020, 2021)):
    rows.append('%s,not available' % iso(d))
write('text-in-nav-column.csv', '\n'.join(rows))

# 5 -- a real PDF. Minimal, but a genuine one: %PDF header, xref, trailer.
def pdf(path, line):
    objs = [
        b'<< /Type /Catalog /Pages 2 0 R >>',
        b'<< /Type /Pages /Kids [3 0 R] /Count 1 >>',
        b'<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] '
        b'/Resources << /Font << /F1 5 0 R >> >> /Contents 4 0 R >>',
        None,
        b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>',
    ]
    stream = ('BT /F1 12 Tf 72 760 Td (%s) Tj ET' % line).encode()
    objs[3] = b'<< /Length %d >>\nstream\n%s\nendstream' % (len(stream), stream)
    out, offsets = b'%PDF-1.7\n%\xe2\xe3\xcf\xd3\n', []
    for i, body in enumerate(objs, start=1):
        offsets.append(len(out))
        out += b'%d 0 obj\n' % i + body + b'\nendobj\n'
    start = len(out)
    out += b'xref\n0 %d\n0000000000 65535 f \n' % (len(objs) + 1)
    for off in offsets:
        out += b'%010d 00000 n \n' % off
    out += (b'trailer\n<< /Size %d /Root 1 0 R >>\nstartxref\n%d\n%%%%EOF\n'
            % (len(objs) + 1, start))
    with open(path, 'wb') as f:
        f.write(out)
    return path

pdf(os.path.join(OUT, 'account-statement.pdf'), 'Consolidated Account Statement')
# the same bytes with a .csv name, to prove the check reads contents not endings
pdf(os.path.join(OUT, 'statement-renamed.csv'), 'Consolidated Account Statement')

# 6 -- Excel. Three junk rows above the header, dates as real Excel dates.
try:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'NAV History'
    ws.append(['Alpha Asset Management'])
    ws.append(['Scheme NAV history report'])
    ws.append([])
    ws.append(['NAV Date', 'Scheme', 'NAV (Rs.)'])
    v = 50.0
    for d in days(2012, 2025):
        ws.append([d, 'Alpha Bluechip Fund - Direct Growth', round(v, 4)])
        v *= (1.13) ** (1 / 365.2425)
    for c in ws['A']:
        c.number_format = 'DD-MMM-YYYY'
    wb.save(os.path.join(OUT, 'amc-nav-history.xlsx'))

    # and one holding three schemes, to drive the picker from an .xlsx
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = 'All schemes'
    ws2.append(['Date', 'Scheme Name', 'NAV'])
    for name, rate in [('Alpha Fund - Direct Growth', 0.14),
                       ('Alpha Fund - Regular Growth', 0.11),
                       ('Beta Fund - Direct Growth', 0.08)]:
        v = 10.0
        for d in days(2012, 2025):
            ws2.append([d, name, round(v, 4)])
            v *= (1 + rate) ** (1 / 365.2425)
    for c in ws2['A']:
        c.number_format = 'DD-MMM-YYYY'
    wb2.save(os.path.join(OUT, 'amc-three-schemes.xlsx'))
except ImportError:
    print('openpyxl missing: the two .xlsx fixtures were not written', file=sys.stderr)

print(OUT)
