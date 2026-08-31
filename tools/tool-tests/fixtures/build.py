"""Rebuild three-tab-statement.xlsx. See README.md for why it is a file."""
import os
from openpyxl import Workbook

HERE = os.path.dirname(os.path.abspath(__file__))
wb = Workbook()
wb.active.title = "Cover"
wb.active["A1"] = "Consolidated Account Statement"
summary = wb.create_sheet("Summary")
summary["A1"], summary["A2"] = "Folio", "123"
txns = wb.create_sheet("Transaction Details")
txns.append(["Date", "Transaction Type", "Amount"])
txns.append(["2024-01-01", "Purchase", 100000])
txns.append(["2024-06-01", "SIP", 5000])
txns.append(["2025-01-01", "Switch Out", 20000])
wb.save(os.path.join(HERE, "three-tab-statement.xlsx"))
print("wrote three-tab-statement.xlsx")
